import re
import tempfile
import zipfile
from pathlib import Path
from typing import List, Tuple

import fitz  # PyMuPDF
from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook

app = Flask(__name__)

# Limit upload size (100MB)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024


# ----------------------------
# Helper Functions
# ----------------------------
def sanitize_filename(name: str) -> str:
    name = str(name).strip()
    name = re.sub(r'[\\/*?:"<>|]', "_", name)
    return name.rstrip(". ") or "output"


def parse_page_range(page_range_value) -> Tuple[int, int]:
    text = str(page_range_value).strip().lower()
    text = text.replace("–", "-").replace("—", "-")

    numbers = [int(x) for x in re.findall(r"\d+", text)]

    if not numbers:
        raise ValueError(f"Invalid page range: {page_range_value}")

    if len(numbers) == 1:
        return numbers[0], numbers[0]

    return numbers[0], numbers[1]


def read_split_config(excel_path: Path) -> List[dict]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    configs = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        page_range, output_format, output_name = row[:3]

        if not page_range or not output_format or not output_name:
            continue

        start, end = parse_page_range(page_range)

        configs.append({
            "start": start,
            "end": end,
            "format": str(output_format).lower(),
            "name": sanitize_filename(output_name)
        })

    return configs


def save_pdf(doc, start, end, path):
    new_doc = fitz.open()
    new_doc.insert_pdf(doc, from_page=start - 1, to_page=end - 1)
    new_doc.save(path)
    new_doc.close()


def save_jpg(doc, start, end, base_path):
    for i, page_no in enumerate(range(start, end + 1), start=1):
        page = doc.load_page(page_no - 1)
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))

        if start == end:
            path = base_path.with_suffix(".jpg")
        else:
            path = base_path.parent / f"{base_path.stem}_{i}.jpg"

        pix.save(path)


# ----------------------------
# Main Process
# ----------------------------
def process(pdf_path, excel_path, output_dir):
    config = read_split_config(excel_path)

    doc = fitz.open(pdf_path)

    for item in config:
        start = item["start"]
        end = item["end"]
        fmt = item["format"]
        name = item["name"]

        output_base = output_dir / name

        if fmt == "pdf":
            save_pdf(doc, start, end, output_base.with_suffix(".pdf"))

        elif fmt in ["jpg", "jpeg"]:
            save_jpg(doc, start, end, output_base)

    doc.close()


def make_zip(folder, zip_path):
    with zipfile.ZipFile(zip_path, "w") as z:
        for file in folder.rglob("*"):
            if file.is_file():
                z.write(file, file.relative_to(folder))


# ----------------------------
# Routes
# ----------------------------
@app.route("/")
def home():
    return render_template("index.html")


@app.route("/process", methods=["POST"])
def run():
    pdf = request.files.get("pdf")
    excel = request.files.get("excel")

    if not pdf or not excel:
        return "Upload both files"

    temp = Path(tempfile.mkdtemp())

    pdf_path = temp / pdf.filename
    excel_path = temp / excel.filename
    output = temp / "output"

    output.mkdir()

    pdf.save(pdf_path)
    excel.save(excel_path)

    process(pdf_path, excel_path, output)

    zip_path = temp / "result.zip"
    make_zip(output, zip_path)

    return send_file(zip_path, as_attachment=True)


# ----------------------------
# Run App
# ----------------------------
if __name__ == "__main__":
    app.run(debug=True)